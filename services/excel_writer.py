# services/excel_writer.py
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import numpy as np
import os

CELL_MAP = {
    "date": "B4",
    "customer": "B3",
    "total_daily": "E10",
    "total_mtd": "E11",
    "total_ytd": "E12",
    "plf_percent": "E13",
    "breakdown_hours": "C20",
    "weather": "C21",
    "generation_hours": "C22",
    "operating_hours": "C23",
}

def _to_scalar(v):
    if v is None:
        return None
    if isinstance(v, (np.generic,)):
        return v.item()
    if isinstance(v, (pd.Series, pd.Index)) and v.size == 1:
        return v.iloc[0]
    if isinstance(v, (pd.Series, pd.Index, pd.DataFrame, list, tuple, dict, set)):
        return str(v)
    if isinstance(v, (str, float, int, bool)):
        return v
    try:
        return float(v)
    except Exception:
        return str(v)

def _top_left_if_merged(ws: Worksheet, coord: str) -> str:
    """If coord lies inside a merged range, return that range’s top-left coordinate."""
    try:
        col_letters, row = coordinate_from_string(coord)
        col = column_index_from_string(col_letters)
    except Exception:
        return coord  # if invalid, just return as-is

    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = rng.bounds
        if (min_col <= col <= max_col) and (min_row <= row <= max_row):
            # Redirect to top-left of merged block
            return f"{get_column_letter(min_col)}{min_row}"
    return coord

def write_report_from_template(template_path, out_path, context, inverter_rows):
    """
    context: dict with keys from CELL_MAP
    inverter_rows: list of tuples (name, daily_kwh, monthly_kwh)
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    # data_only=False keeps formulas; header/footer warnings can be ignored
    wb = load_workbook(template_path, data_only=False)
    ws = wb.active

    # Fill header / KPI cells (handle merged cells)
    for key, coord in CELL_MAP.items():
        safe_coord = _top_left_if_merged(ws, coord)
        ws[safe_coord].value = _to_scalar(context.get(key, ""))

    # Write inverter table (adjust start row to match your template)
    start_row = 30
    for i, (name, dval, mval) in enumerate(inverter_rows):
        r = start_row + i
        ws[f"A{r}"].value = _to_scalar(name)
        ws[f"B{r}"].value = _to_scalar(dval)
        ws[f"C{r}"].value = _to_scalar(mval)

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path
